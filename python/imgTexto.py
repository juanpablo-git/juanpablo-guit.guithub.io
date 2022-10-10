from csv import list_dialects
import pytesseract
import cv2
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

#ler o exel desejado
wb = load_workbook("puy.xlsx")
ws = wb.active
licalizacao = None

# pegar a ultima linha escrita
licalizacao = len(ws["A"])

#pegar a imagem
imagem = cv2.imread("codBarra.png")
texto = pytesseract.image_to_string(imagem)

#tranformar o texto da imagem em array
arryText = texto.split("\n")
ws = wb.active


for t in arryText:
    if t != ' ' and t != '' and t != "\n":
         print(t)
         ws.append([t])
         wb.save("puy.xlsx")
